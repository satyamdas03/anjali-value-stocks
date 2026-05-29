"""
Build US Stock Analysis Excel with corrected color hierarchies matching sample data patterns.
Also builds SmallCap 600 + MidCap 400 sheet with index-relative coloring.

Color hierarchies (decoded from sample data, verified against 249 Indian stocks):
  Valuation (PE, Future PE, PEG, Future PEG): Q1=White, Q2=DG, Q3=LG, Q4=LR, Q5=DR. NaN=DR.
    Ultra-cheap = value trap concern (White). Sweet spot in Q2 (DG).
  Risk (Std, Beta): Q1=LR, Q2=White, Q3=LG, Q4=DG, Q5=DR.
    Too safe = missed returns (LR). Sweet spot in Q4 (DG).
  Returns (3M, 6M, 1Yr, 2Yr): Q1=DR, Q2=LR, Q3=White, Q4=LG, Q5=DG. Higher=better.
  Growth (6 columns): Q1=DR, Q2=LR, Q3=White, Q4=LG, Q5=DG. Loss-making profit cols=DR.

Scores: DG=+1, LG=+0.5, White=0, LR=-0.5, DR=-1. Sum per category.
GROWTH SCORE uses only 4 cols (excludes QoQ).
Ratios + Size columns: uncolored.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time
import warnings
import requests
import yfinance as yf

warnings.filterwarnings("ignore")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(SCRIPT_DIR, "US_Stock_Data.csv")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "US_Stock_Analysis_Coloured.xlsx")

# --- Fills ---
FILLS = {
    "DG": PatternFill(start_color="34CF58", end_color="34CF58", fill_type="solid"),
    "LG": PatternFill(start_color="99F2BB", end_color="99F2BB", fill_type="solid"),
    "White": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "LR": PatternFill(start_color="E3CACA", end_color="E3CACA", fill_type="solid"),
    "DR": PatternFill(start_color="F5AEAE", end_color="F5AEAE", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SCORE_MAP = {"DG": 1.0, "LG": 0.5, "White": 0.0, "LR": -0.5, "DR": -1.0}

# --- Column Definitions ---
# (col_key, header_name, format_str)
ALL_COLUMNS = [
    # Core (no color)
    ("Ticker", "Ticker", None),
    ("Sector", "Sector", None),
    ("Sub Sector", "Sub Sector", None),
    # Growth (6 columns)
    ("Sales YoY Growth", "Sales YoY Growth", "0.00"),
    ("NetProfit YoY Growth", "NetProfit YoY Growth", "0.00"),
    ("Sales TTM 1Yr Growth", "Sales TTM 1Yr Growth", "0.00"),
    ("NetProfit TTM 1Yr Growth", "NetProfit TTM 1Yr Growth", "0.00"),
    ("QoQ Sales Growth", "QoQ Sales Growth", "0.00"),
    ("QoQ Profit Growth", "QoQ Profit Growth", "0.00"),
    # Returns (4 columns)
    ("3M Return", "3M Return", "0.00"),
    ("6M Return", "6M Return", "0.00"),
    ("1Yr Return", "1Yr Return", "0.00"),
    ("2Yr Return", "2Yr Return", "0.00"),
    # Valuation (4 columns)
    ("PE Ratio", "PE Ratio", "0.00"),
    ("Future PE", "Future PE", "0.00"),
    ("TTM PEG", "TTM PEG", "0.00"),
    ("Future PEG", "Future PEG", "0.00"),
    # Ratios (3 columns, UNCOLORED)
    ("PB Ratio", "PB Ratio", "0.00"),
    ("EV/Sales", "EV/Sales", "0.00"),
    ("EV/EBITDA", "EV/EBITDA", "0.00"),
    # Size (3 columns, UNCOLORED)
    ("Market Cap (Billions)", "Market Cap (B)", "0.00"),
    ("Revenue (Billions)", "Revenue (B)", "0.00"),
    ("TTM Revenue (Billions)", "TTM Revenue (B)", "0.00"),
    # Risk (4 columns)
    ("QtrStd", "QtrStd", "0.00"),
    ("YrStd", "YrStd", "0.00"),
    ("Qtr Beta", "Qtr Beta", "0.0000"),
    ("Yr Beta", "Yr Beta", "0.0000"),
    # Scores (4 columns, no color)
    ("RETURN SCORE", "RETURN SCORE", "0.00"),
    ("GROWTH SCORE", "GROWTH SCORE", "0.00"),
    ("VALUATION SCORE", "VALUATION SCORE", "0.00"),
    ("RISK SCORE", "RISK SCORE", "0.00"),
]

GROWTH_COLS = [
    "Sales YoY Growth", "NetProfit YoY Growth",
    "Sales TTM 1Yr Growth", "NetProfit TTM 1Yr Growth",
    "QoQ Sales Growth", "QoQ Profit Growth",
]
# GROWTH SCORE only uses 4 cols (excludes QoQ)
GROWTH_SCORE_COLS = [
    "Sales YoY Growth", "NetProfit YoY Growth",
    "Sales TTM 1Yr Growth", "NetProfit TTM 1Yr Growth",
]
RETURN_COLS = ["3M Return", "6M Return", "1Yr Return", "2Yr Return"]
VALUATION_COLS = ["PE Ratio", "Future PE", "TTM PEG", "Future PEG"]
RATIO_COLS = ["PB Ratio", "EV/Sales", "EV/EBITDA"]
SIZE_COLS = ["Market Cap (Billions)", "Revenue (Billions)", "TTM Revenue (Billions)"]
RISK_COLS = ["QtrStd", "YrStd", "Qtr Beta", "Yr Beta"]

# Loss flags -> column mapping
LOSS_FLAG_MAP = {
    "NetProfit YoY Growth": "_Loss_Profit_YoY",
    "NetProfit TTM 1Yr Growth": "_Loss_Profit_TTM",
    "QoQ Profit Growth": "_Loss_Profit_QoQ",
}

UNCOLORED_COLS = RATIO_COLS + SIZE_COLS


def _quantile_bounds(series, n_groups):
    clean = series.dropna()
    if len(clean) < n_groups:
        return None
    return [clean.quantile(i / n_groups) for i in range(1, n_groups)]


# ============================================================
# COLOR FUNCTIONS (corrected per sample data patterns)
# ============================================================

def color_valuation(df, col):
    """Q1(lowest PE)=White, Q2=DG, Q3=LG, Q4=LR, Q5(highest)=DR. NaN=DR.
    Ultra-cheap = value trap (White). Sweet spot in Q2 (DG)."""
    series = pd.to_numeric(df[col], errors="coerce")
    valid = series.dropna()

    if len(valid) < 5:
        return pd.Series("White", index=df.index)

    q20 = valid.quantile(0.20)
    q40 = valid.quantile(0.40)
    q60 = valid.quantile(0.60)
    q80 = valid.quantile(0.80)

    def assign(v):
        if pd.isna(v):
            return "DR"
        if v <= q20:
            return "White"
        if v <= q40:
            return "DG"
        if v <= q60:
            return "LG"
        if v <= q80:
            return "LR"
        return "DR"

    return series.apply(assign)


def color_risk(df, col):
    """Q1(lowest risk)=LR, Q2=White, Q3=LG, Q4=DG, Q5(highest)=DR.
    Too safe = missed returns (LR). Sweet spot in Q4 (DG)."""
    series = pd.to_numeric(df[col], errors="coerce")
    valid = series.dropna()

    if len(valid) < 5:
        return pd.Series("White", index=df.index)

    q20 = valid.quantile(0.20)
    q40 = valid.quantile(0.40)
    q60 = valid.quantile(0.60)
    q80 = valid.quantile(0.80)

    def assign(v):
        if pd.isna(v):
            return "White"
        if v <= q20:
            return "LR"
        if v <= q40:
            return "White"
        if v <= q60:
            return "LG"
        if v <= q80:
            return "DG"
        return "DR"

    return series.apply(assign)


def color_returns(df, col):
    """Standard ascending: Q1(worst)=DR, Q2=LR, Q3=White, Q4=LG, Q5(best)=DG."""
    series = pd.to_numeric(df[col], errors="coerce")
    valid = series.dropna()

    if len(valid) < 5:
        return pd.Series("White", index=df.index)

    q20 = valid.quantile(0.20)
    q40 = valid.quantile(0.40)
    q60 = valid.quantile(0.60)
    q80 = valid.quantile(0.80)

    def assign(v):
        if pd.isna(v):
            return "White"
        if v <= q20:
            return "DR"
        if v <= q40:
            return "LR"
        if v <= q60:
            return "White"
        if v <= q80:
            return "LG"
        return "DG"

    return series.apply(assign)


def color_growth(df, col):
    """Standard ascending: Q1(worst)=DR, Q2=LR, Q3=White, Q4=LG, Q5(best)=DG.
    Loss-making profit columns -> DR. NaN -> White."""
    series = pd.to_numeric(df[col], errors="coerce")
    loss_flag = LOSS_FLAG_MAP.get(col)

    if loss_flag and loss_flag in df.columns:
        loss_mask = df[loss_flag].fillna(False)
    else:
        loss_mask = pd.Series(False, index=df.index)

    profitable = series[~loss_mask].dropna()

    if len(profitable) < 4:
        def fallback(v, is_loss):
            if pd.isna(v):
                return "White"
            return "DR" if is_loss else "White"
        result = pd.Series("White", index=df.index)
        for idx in df.index:
            result.at[idx] = fallback(
                series.at[idx] if idx in series.index else None,
                loss_mask.at[idx] if idx in loss_mask.index else False)
        return result

    q20 = profitable.quantile(0.20)
    q40 = profitable.quantile(0.40)
    q60 = profitable.quantile(0.60)
    q80 = profitable.quantile(0.80)

    def assign(v, is_loss):
        if pd.isna(v):
            return "White"
        if is_loss:
            return "DR"
        if v <= q20:
            return "DR"
        if v <= q40:
            return "LR"
        if v <= q60:
            return "White"
        if v <= q80:
            return "LG"
        return "DG"

    result = pd.Series("White", index=df.index)
    for idx in df.index:
        result.at[idx] = assign(
            series.at[idx] if idx in series.index else None,
            loss_mask.at[idx] if idx in loss_mask.index else False)
    return result


# ============================================================
# DATA COLLECTION HELPERS (for Small/Mid cap)
# ============================================================

REVENUE_ROWS = ["TotalRevenue", "Total Revenue", "Net Revenue"]
INCOME_ROWS = [
    "NetIncome", "Net Income", "Net Income Common Stockholders",
    "Net Income Continuous Operations",
]


def safe_round(val, decimals=2):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    return round(float(val), decimals)


def try_financials_row(fin, row_names, col_name):
    row = next((r for r in row_names if r in fin.index), None)
    if row:
        val = fin.loc[row, col_name]
        if pd.notna(val):
            return val
    return None


def fetch_wikipedia_tickers(url):
    """Fetch ticker symbols from a Wikipedia S&P index page."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        tables = pd.read_html(resp.text)
        df = tables[0]
        tickers = df["Symbol"].str.replace(".", "-", regex=False).tolist()
        return list(dict.fromkeys(tickers))
    except Exception as e:
        print(f"  Wikipedia fetch failed: {e}")
        return []


def collect_index_data(tickers, spy_hist_3mo, spy_hist_1y, index_label):
    """Collect yfinance data for a list of tickers. Returns DataFrame."""
    spy_qtr_returns = spy_hist_3mo["Close"].pct_change().dropna()
    spy_yr_returns = spy_hist_1y["Close"].pct_change().dropna()

    rows = []
    total = len(tickers)

    for i, ticker in enumerate(tickers):
        row = {col: None for col in [c[0] for c in ALL_COLUMNS]}
        row["Ticker"] = ticker
        row["Index"] = index_label

        try:
            stock = yf.Ticker(ticker)
            info = stock.info or {}

            row["Sector"] = info.get("sector", "")
            row["Sub Sector"] = info.get("industry", "")

            # Price History & Returns
            hist_3mo = stock.history(period="3mo")
            hist_6mo = stock.history(period="6mo")
            hist_1y = stock.history(period="1y")
            hist_2y = stock.history(period="2y")

            if len(hist_1y) > 1:
                price_now = hist_1y["Close"].iloc[-1]
                price_3m = hist_3mo["Close"].iloc[0] if len(hist_3mo) > 0 else None
                price_6m = hist_6mo["Close"].iloc[0] if len(hist_6mo) > 0 else None
                price_1y = hist_1y["Close"].iloc[0]
                price_2y = hist_2y["Close"].iloc[0] if len(hist_2y) > 0 else None

                row["3M Return"] = safe_round(((price_now / price_3m) - 1) * 100) if price_3m else None
                row["6M Return"] = safe_round(((price_now / price_6m) - 1) * 100) if price_6m else None
                row["1Yr Return"] = safe_round(((price_now / price_1y) - 1) * 100) if price_1y else None
                row["2Yr Return"] = safe_round(((price_now / price_2y) - 1) * 100) if price_2y else None

            # Risk Metrics
            if len(hist_1y) > 20:
                daily_returns = hist_1y["Close"].pct_change().dropna()
                qtr_daily = daily_returns.tail(63)
                row["QtrStd"] = safe_round(qtr_daily.std() * np.sqrt(252) * 100)
                row["YrStd"] = safe_round(daily_returns.std() * np.sqrt(252) * 100)

            # Beta
            if len(hist_3mo) > 10 and len(spy_qtr_returns) > 10:
                stock_qtr_ret = hist_3mo["Close"].pct_change().dropna()
                common = stock_qtr_ret.index.intersection(spy_qtr_returns.index)
                if len(common) > 10:
                    cov = np.cov(stock_qtr_ret.loc[common], spy_qtr_returns.loc[common])[0][1]
                    var = np.var(spy_qtr_returns.loc[common])
                    row["Qtr Beta"] = safe_round(cov / var, 4) if var > 0 else None

            if len(hist_1y) > 20 and len(spy_yr_returns) > 20:
                stock_yr_ret = hist_1y["Close"].pct_change().dropna()
                common = stock_yr_ret.index.intersection(spy_yr_returns.index)
                if len(common) > 20:
                    cov = np.cov(stock_yr_ret.loc[common], spy_yr_returns.loc[common])[0][1]
                    var = np.var(spy_yr_returns.loc[common])
                    row["Yr Beta"] = safe_round(cov / var, 4) if var > 0 else None

            # Valuation
            pe = info.get("trailingPE")
            pb = info.get("priceToBook")
            row["PE Ratio"] = safe_round(pe) if pe else None
            row["PB Ratio"] = safe_round(pb) if pb and pb > 0 else None
            row["EV/Sales"] = safe_round(info.get("enterpriseToRevenue"))
            row["EV/EBITDA"] = safe_round(info.get("enterpriseToEbitda"))

            # Size
            mc = info.get("marketCap")
            row["Market Cap (Billions)"] = safe_round(mc / 1e9) if mc else None
            rev = info.get("totalRevenue")
            row["Revenue (Billions)"] = safe_round(rev / 1e9) if rev else None

            # Growth from .info
            rev_g = info.get("revenueGrowth")
            earn_g = info.get("earningsGrowth")
            row["Sales YoY Growth"] = safe_round(rev_g * 100) if rev_g is not None else None
            row["NetProfit YoY Growth"] = safe_round(earn_g * 100) if earn_g is not None else None

            # TTM Growth from Annual Financials
            fin = stock.financials
            if not fin.empty and fin.shape[1] >= 2:
                latest_col = fin.columns[0]
                prior_col = fin.columns[1]

                rev_curr = try_financials_row(fin, REVENUE_ROWS, latest_col)
                rev_prev = try_financials_row(fin, REVENUE_ROWS, prior_col)
                if rev_curr and rev_prev and rev_prev != 0:
                    row["Sales TTM 1Yr Growth"] = safe_round((rev_curr / rev_prev - 1) * 100, 1)
                if rev_curr:
                    row["TTM Revenue (Billions)"] = safe_round(rev_curr / 1e9)

                ni_curr = try_financials_row(fin, INCOME_ROWS, latest_col)
                ni_prev = try_financials_row(fin, INCOME_ROWS, prior_col)
                if ni_curr and ni_prev and ni_prev != 0:
                    row["NetProfit TTM 1Yr Growth"] = safe_round((ni_curr / ni_prev - 1) * 100, 1)

            # QoQ Growth
            qfin = stock.quarterly_financials
            if not qfin.empty and qfin.shape[1] >= 2:
                qlatest = qfin.columns[0]
                qprior = qfin.columns[1]

                qrev_curr = try_financials_row(qfin, REVENUE_ROWS, qlatest)
                qrev_prev = try_financials_row(qfin, REVENUE_ROWS, qprior)
                if qrev_curr and qrev_prev and qrev_prev != 0:
                    row["QoQ Sales Growth"] = safe_round((qrev_curr / qrev_prev - 1) * 100, 1)
                elif row["QoQ Sales Growth"] is None:
                    rqg = info.get("revenueQuarterlyGrowth")
                    row["QoQ Sales Growth"] = safe_round(rqg * 100) if rqg is not None else None

                qni_curr = try_financials_row(qfin, INCOME_ROWS, qlatest)
                qni_prev = try_financials_row(qfin, INCOME_ROWS, qprior)
                if qni_curr and qni_prev and qni_prev != 0:
                    row["QoQ Profit Growth"] = safe_round((qni_curr / qni_prev - 1) * 100, 1)
                elif row["QoQ Profit Growth"] is None:
                    eqg = info.get("earningsQuarterlyGrowth")
                    row["QoQ Profit Growth"] = safe_round(eqg * 100) if eqg is not None else None

            # Cap extreme QoQ Profit
            if row["QoQ Profit Growth"] is not None:
                if row["QoQ Profit Growth"] > 500:
                    row["QoQ Profit Growth"] = 500.0
                elif row["QoQ Profit Growth"] < -500:
                    row["QoQ Profit Growth"] = -500.0

            # Future PE + PEG
            if row["PE Ratio"] is not None and row["NetProfit TTM 1Yr Growth"] is not None:
                row["Future PE"] = safe_round(row["PE Ratio"] * (1 + row["NetProfit TTM 1Yr Growth"] / 100))
            if row["PE Ratio"] is not None and row["NetProfit TTM 1Yr Growth"] is not None and row["NetProfit TTM 1Yr Growth"] != 0:
                row["TTM PEG"] = safe_round(row["PE Ratio"] / row["NetProfit TTM 1Yr Growth"])
            if row["Future PE"] is not None and row["NetProfit TTM 1Yr Growth"] is not None and row["NetProfit TTM 1Yr Growth"] != 0:
                row["Future PEG"] = safe_round(row["Future PE"] / row["NetProfit TTM 1Yr Growth"])

            # Loss Flags
            row["_Loss_Profit_YoY"] = (row["NetProfit YoY Growth"] is not None and row["NetProfit YoY Growth"] < 0)
            row["_Loss_Profit_TTM"] = (row["NetProfit TTM 1Yr Growth"] is not None and row["NetProfit TTM 1Yr Growth"] < 0)
            row["_Loss_Profit_QoQ"] = (row["QoQ Profit Growth"] is not None and row["QoQ Profit Growth"] < 0)

        except Exception as e:
            print(f"  [{i+1}/{total}] {ticker}: ERROR — {e}")

        rows.append(row)

        if (i + 1) % 25 == 0 or i == total - 1:
            print(f"  [{i+1}/{total}] {ticker}: PE={row.get('PE Ratio')}, "
                  f"1Yr={row.get('1Yr Return')}  ({(i+1)/total*100:.0f}%)")

        time.sleep(0.3)

    return pd.DataFrame(rows)


# ============================================================
# EXCEL BUILDING
# ============================================================

def assign_colors(df):
    """Assign colors for all columns. Returns color DataFrame."""
    color_df = pd.DataFrame(index=df.index)
    for col in GROWTH_COLS:
        if col in df.columns:
            color_df[col] = color_growth(df, col)
    for col in RETURN_COLS:
        if col in df.columns:
            color_df[col] = color_returns(df, col)
    for col in VALUATION_COLS:
        if col in df.columns:
            color_df[col] = color_valuation(df, col)
    for col in RISK_COLS:
        if col in df.columns:
            color_df[col] = color_risk(df, col)
    return color_df


def compute_scores(df, color_df):
    """Compute sum-based scores from color assignments."""
    def sum_scores(r, col_list):
        total = 0.0
        for c in col_list:
            if c in color_df.columns and pd.notna(color_df.at[r.name, c]):
                total += SCORE_MAP.get(color_df.at[r.name, c], 0)
        return round(total, 2)

    scores = pd.DataFrame(index=df.index)
    scores["RETURN SCORE"] = df.apply(lambda r: sum_scores(r, RETURN_COLS), axis=1)
    scores["GROWTH SCORE"] = df.apply(lambda r: sum_scores(r, GROWTH_SCORE_COLS), axis=1)
    scores["VALUATION SCORE"] = df.apply(lambda r: sum_scores(r, VALUATION_COLS), axis=1)
    scores["RISK SCORE"] = df.apply(lambda r: sum_scores(r, RISK_COLS), axis=1)
    return scores


def write_sheet(ws, df, color_df, scores, index_label=None):
    """Write data rows to an openpyxl worksheet with color coding."""
    headers = [h for _, h, _ in ALL_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = BORDER

    for row_idx, (df_idx, row) in enumerate(df.iterrows(), 2):
        for col_idx, (col_key, header, fmt) in enumerate(ALL_COLUMNS, 1):
            if col_key in scores.columns:
                value = scores.at[df_idx, col_key]
            elif col_key in df.columns:
                value = row.get(col_key)
            else:
                value = None

            if pd.isna(value) or value is None:
                cell = ws.cell(row=row_idx, column=col_idx, value=None)
            elif fmt and isinstance(value, (int, float, np.floating)):
                cell = ws.cell(row=row_idx, column=col_idx, value=round(float(value), 4))
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx > 3 else "left")

            if fmt and isinstance(value, (int, float, np.floating)) and not pd.isna(value):
                cell.number_format = fmt

            if col_key in color_df.columns and pd.notna(color_df.at[df_idx, col_key]):
                color = color_df.at[df_idx, col_key]
                if color in FILLS:
                    cell.fill = FILLS[color]


def set_column_widths(ws):
    col_widths = {
        "Ticker": 10, "Sector": 25, "Sub Sector": 35,
        "Sales YoY Growth": 14, "NetProfit YoY Growth": 14,
        "Sales TTM 1Yr Growth": 14, "NetProfit TTM 1Yr Growth": 14,
        "QoQ Sales Growth": 14, "QoQ Profit Growth": 14,
        "3M Return": 12, "6M Return": 12, "1Yr Return": 12, "2Yr Return": 12,
        "PE Ratio": 12, "Future PE": 12, "TTM PEG": 12, "Future PEG": 12,
        "PB Ratio": 12, "EV/Sales": 12, "EV/EBITDA": 12,
        "Market Cap (B)": 16, "Revenue (B)": 16, "TTM Revenue (B)": 16,
        "QtrStd": 10, "YrStd": 10, "Qtr Beta": 14, "Yr Beta": 14,
        "RETURN SCORE": 14, "GROWTH SCORE": 14,
        "VALUATION SCORE": 14, "RISK SCORE": 14,
    }
    for col_idx, (col_key, header, _) in enumerate(ALL_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 13)


def add_legend(ws, start_row):
    """Add color legend below data."""
    sr = start_row
    ws.cell(row=sr, column=1, value="COLOR LEGEND").font = Font(bold=True, size=11)
    ws.cell(row=sr + 1, column=1, value="Dark Green (+1)").fill = FILLS["DG"]
    ws.cell(row=sr + 1, column=2, value="Top tier — strongest positive signal")
    ws.cell(row=sr + 2, column=1, value="Light Green (+0.5)").fill = FILLS["LG"]
    ws.cell(row=sr + 2, column=2, value="Above average")
    ws.cell(row=sr + 3, column=1, value="White (0)").fill = FILLS["White"]
    ws.cell(row=sr + 3, column=2, value="Neutral / average / value trap zone")
    ws.cell(row=sr + 4, column=1, value="Light Red (-0.5)").fill = FILLS["LR"]
    ws.cell(row=sr + 4, column=2, value="Below average")
    ws.cell(row=sr + 5, column=1, value="Dark Red (-1)").fill = FILLS["DR"]
    ws.cell(row=sr + 5, column=2, value="Bottom tier — weakest signal")

    ws.cell(row=sr + 7, column=1, value="COLOR HIERARCHIES (matched to sample data):").font = Font(bold=True, size=10)
    ws.cell(row=sr + 8, column=1, value="Growth: Q1(worst)=DR, Q2=LR, Q3=White, Q4=LG, Q5(best)=DG. Loss-making profit cols=DR.")
    ws.cell(row=sr + 9, column=1, value="Returns: Q1(worst)=DR, Q2=LR, Q3=White, Q4=LG, Q5(best)=DG. Higher return = better.")
    ws.cell(row=sr + 10, column=1, value="Valuation: Q1(cheapest)=White(value trap), Q2=DG(sweet spot), Q3=LG, Q4=LR, Q5(most expensive)=DR. NaN=DR.")
    ws.cell(row=sr + 11, column=1, value="Risk: Q1(safest)=LR(missed returns), Q2=White, Q3=LG, Q4=DG(sweet spot), Q5(riskiest)=DR.")
    ws.cell(row=sr + 12, column=1, value="Ratios (PB, EV/Sales, EV/EBITDA) + Size (Market Cap, Revenue, TTM Revenue): UNCOLORED.")

    ws.cell(row=sr + 14, column=1, value="SCORING:").font = Font(bold=True, size=10)
    ws.cell(row=sr + 15, column=1, value="DG=+1, LG=+0.5, White=0, LR=-0.5, DR=-1. Sum across category columns.")
    ws.cell(row=sr + 16, column=1, value="GROWTH SCORE uses 4 cols (Sales YoY, NetProfit YoY, Sales TTM, NetProfit TTM) — QoQ EXCLUDED.")
    ws.cell(row=sr + 17, column=1, value="Source: Yahoo Finance (yfinance). Benchmark: S&P 500 (SPY) for Beta.")
    ws.cell(row=sr + 18, column=1, value=f"Future PE = Current PE x (1 + TTM Profit Growth/100)")


# ============================================================
# MAIN S&P 500 SHEET
# ============================================================

def build_sp500_sheet(wb):
    """Build S&P 500 sheet from US_Stock_Data.csv."""
    df = pd.read_csv(CSV_PATH)
    df = df.drop_duplicates(subset=["Ticker"])
    df = df.sort_values("Ticker").reset_index(drop=True)

    loss_cols = [c for c in df.columns if c.startswith("_Loss_")]
    df_clean = df.drop(columns=loss_cols, errors="ignore")

    print(f"Building S&P 500 sheet: {len(df_clean)} stocks, {len(ALL_COLUMNS)} columns")

    color_df = assign_colors(df)
    scores = compute_scores(df, color_df)

    ws = wb.create_sheet("S&P 500 Analysis")
    write_sheet(ws, df_clean, color_df, scores)
    set_column_widths(ws)
    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(ALL_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(df_clean) + 1}"
    add_legend(ws, len(df_clean) + 3)

    print("=== S&P 500 Score Ranges ===")
    for sc in ["RETURN SCORE", "GROWTH SCORE", "VALUATION SCORE", "RISK SCORE"]:
        vals = scores[sc]
        print(f"  {sc}: min={vals.min():.1f}, max={vals.max():.1f}, median={vals.median():.1f}")

    return scores


# ============================================================
# SMALL CAP 600 + MID CAP 400 SHEET
# ============================================================

def build_small_mid_sheet(wb):
    """Build SmallCap 600 + MidCap 400 sheet with index-relative coloring."""
    print("\n=== Building SmallCap 600 + MidCap 400 Sheet ===\n")

    # Fetch ticker lists
    mid_url = "https://en.wikipedia.org/wiki/List_of_S%26P_400_companies"
    small_url = "https://en.wikipedia.org/wiki/List_of_S%26P_600_companies"

    print("Fetching MidCap 400 tickers...")
    mid_tickers = fetch_wikipedia_tickers(mid_url)
    print(f"  Got {len(mid_tickers)} MidCap tickers")

    print("Fetching SmallCap 600 tickers...")
    small_tickers = fetch_wikipedia_tickers(small_url)
    print(f"  Got {len(small_tickers)} SmallCap tickers")

    # Pre-fetch SPY
    print("Fetching SPY benchmark...")
    spy = yf.Ticker("SPY")
    spy_hist_3mo = spy.history(period="3mo")
    spy_hist_1y = spy.history(period="1y")

    # Collect data
    print(f"\nCollecting MidCap 400 data ({len(mid_tickers)} stocks)...")
    mid_df = collect_index_data(mid_tickers, spy_hist_3mo, spy_hist_1y, "MidCap 400")
    print(f"\nCollecting SmallCap 600 data ({len(small_tickers)} stocks)...")
    small_df = collect_index_data(small_tickers, spy_hist_3mo, spy_hist_1y, "SmallCap 600")

    # Combine
    combined = pd.concat([mid_df, small_df], ignore_index=True)
    combined = combined.drop_duplicates(subset=["Ticker"])
    combined = combined.sort_values(["Index", "Ticker"]).reset_index(drop=True)

    print(f"\nCombined: {len(combined)} stocks ({len(mid_df)} MidCap + {len(small_df)} SmallCap)")

    # Index-relative coloring: color each index group separately, then merge
    all_colors = []
    all_scores = []

    for idx_label, idx_df in combined.groupby("Index"):
        idx_df = idx_df.reset_index(drop=True)
        print(f"\n--- {idx_label}: {len(idx_df)} stocks ---")

        # Assign colors within index group
        loss_cols_local = [c for c in idx_df.columns if c.startswith("_Loss_")]
        color_idx = assign_colors(idx_df)

        # Add index label back to color for tracking
        color_idx["_idx"] = idx_df["Index"].values

        scores_idx = compute_scores(idx_df, color_idx)

        all_colors.append(color_idx)
        all_scores.append(scores_idx)

        for sc in ["RETURN SCORE", "GROWTH SCORE", "VALUATION SCORE", "RISK SCORE"]:
            vals = scores_idx[sc]
            print(f"  {sc}: min={vals.min():.1f}, max={vals.max():.1f}, median={vals.median():.1f}")

    final_colors = pd.concat(all_colors, ignore_index=True)
    final_scores = pd.concat(all_scores, ignore_index=True)

    # Clean up for display
    combined_display = combined.drop(columns=[c for c in combined.columns if c.startswith("_Loss_")], errors="ignore")

    # Write sheet
    ws = wb.create_sheet("SmallMidCap Analysis")
    write_sheet(ws, combined_display, final_colors, final_scores)
    set_column_widths(ws)
    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(ALL_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(combined_display) + 1}"
    add_legend(ws, len(combined_display) + 3)

    print(f"\nSmallMidCap sheet complete: {len(combined)} stocks")
    return final_scores


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    import sys
    sp500_only = "--sp500-only" in sys.argv
    smallmid_only = "--smallmid-only" in sys.argv

    if smallmid_only:
        from openpyxl import load_workbook
        wb = load_workbook(OUTPUT_PATH)
        build_small_mid_sheet(wb)
        wb.save(OUTPUT_PATH)
        print(f"\nSaved Excel to {OUTPUT_PATH}")
        print("Sheets:", wb.sheetnames)
    else:
        wb = Workbook()
        wb.remove(wb.active)

        build_sp500_sheet(wb)

        if not sp500_only:
            build_small_mid_sheet(wb)
        else:
            print("\nSkipping SmallMidCap sheet (--sp500-only flag set)")

        wb.save(OUTPUT_PATH)
        print(f"\nSaved Excel to {OUTPUT_PATH}")
        print("Sheets:", wb.sheetnames)
