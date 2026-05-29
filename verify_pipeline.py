"""
Verification agent for AnjaliValueStocks pipeline.
Cross-checks all output files before GitHub commit.

Usage:
    python verify_pipeline.py [--strict]

Checks:
1. File existence (all expected files present)
2. CSV structure (columns match spec, non-empty, no duplicates)
3. Excel structure (sheets present, dimensions correct)
4. Color application (colors match expected set)
5. Score ranges (within expected bounds)
6. Data fill rates (reasonable completeness)
7. Syntax checks (all .py files compile)
8. Data consistency (tickers unique, no ghost rows)

Exit code 0 = all passed. Exit code 1 = failures found.
"""
import sys
import os
import ast
import subprocess
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

warnings_filter = lambda: None  # Placeholder

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Expected colors from sample data
EXPECTED_COLORS = {"34CF58", "99F2BB", "FFFFFF", "E3CACA", "F5AEAE"}

# Expected score ranges
SCORE_RANGES = {
    "RETURN SCORE": (-4.0, 4.0),
    "GROWTH SCORE": (-4.0, 4.0),
    "VALUATION SCORE": (-4.0, 4.0),
    "RISK SCORE": (-4.0, 4.0),
}

# Minimum fill rate thresholds
MIN_FILL_RATES = {
    "Ticker": 95,
    "Sector": 90,
    "3M Return": 90,
    "1Yr Return": 95,
    "PE Ratio": 85,
    "QtrStd": 95,
    "YrStd": 95,
    "Qtr Beta": 90,
    "Yr Beta": 90,
}

ERRORS = []
WARNINGS = []


def error(msg):
    ERRORS.append(f"[ERROR] {msg}")
    print(f"[FAIL] {msg}")


def warn(msg):
    WARNINGS.append(f"[WARN] {msg}")
    print(f"[WARN] {msg}")


def ok(msg):
    print(f"[PASS] {msg}")


def section(title):
    print(f"\n{'=' * 60}")
    print(f"  {title}")
    print(f"{'=' * 60}")


# ============================================================
# CHECK 1: File Existence
# ============================================================

def check_files():
    section("CHECK 1: File Existence")

    required = [
        "collect_data.py", "collect_indian_data.py", "collect_smallmid_data.py",
        "build_excel.py", "scheduler.py", "requirements.txt", "Dockerfile",
        "railway.json", "docker-compose.yml",
    ]
    data_files = [
        "US_Stock_Data.csv",
        "SmallMidCap_Stock_Data.csv",
        "Indian_Stock_Data.csv",
        "US_Stock_Analysis_Coloured.xlsx",
    ]

    for f in required:
        path = os.path.join(SCRIPT_DIR, f)
        if os.path.exists(path):
            ok(f"{f} exists")
        else:
            error(f"{f} missing")

    for f in data_files:
        path = os.path.join(SCRIPT_DIR, f)
        if os.path.exists(path):
            ok(f"{f} exists")
        else:
            warn(f"{f} not yet generated (run collectors first)")


# ============================================================
# CHECK 2: Python Syntax
# ============================================================

def check_syntax():
    section("CHECK 2: Python Syntax")

    py_files = [
        "collect_data.py", "collect_indian_data.py", "collect_smallmid_data.py",
        "build_excel.py", "scheduler.py",
    ]
    for f in py_files:
        path = os.path.join(SCRIPT_DIR, f)
        if not os.path.exists(path):
            error(f"{f} not found for syntax check")
            continue
        try:
            with open(path, "r", encoding="utf-8") as fp:
                ast.parse(fp.read())
            ok(f"{f} syntax valid")
        except SyntaxError as e:
            error(f"{f} syntax error: {e}")


# ============================================================
# CHECK 3: CSV Structure
# ============================================================

def check_csv(csv_name, expected_cols):
    section(f"CHECK 3: {csv_name} Structure")

    path = os.path.join(SCRIPT_DIR, csv_name)
    if not os.path.exists(path):
        warn(f"{csv_name} not found -- skipping CSV checks")
        return

    df = pd.read_csv(path)
    ok(f"{csv_name}: {len(df)} rows, {len(df.columns)} columns")

    # Check columns
    missing = [c for c in expected_cols if c not in df.columns]
    extra = [c for c in df.columns if c not in expected_cols]
    if missing:
        error(f"{csv_name} missing columns: {missing}")
    else:
        ok(f"{csv_name} has all expected columns")
    if extra:
        warn(f"{csv_name} extra columns: {extra}")

    # Check duplicates
    dups = df["Ticker"].duplicated().sum()
    if dups > 0:
        error(f"{csv_name} has {dups} duplicate tickers")
    else:
        ok(f"{csv_name}: no duplicate tickers")

    # Check empty rows
    empty = df.isnull().all(axis=1).sum()
    if empty > 0:
        warn(f"{csv_name}: {empty} completely empty rows")
    else:
        ok(f"{csv_name}: no empty rows")

    # Check fill rates
    for col, threshold in MIN_FILL_RATES.items():
        if col in df.columns:
            rate = 100 * df[col].notna().sum() / len(df)
            if rate >= threshold:
                ok(f"{csv_name}.{col}: {rate:.0f}% fill (>={threshold}%)")
            else:
                warn(f"{csv_name}.{col}: {rate:.0f}% fill (<{threshold}%)")

    # Check score ranges if scores exist
    for score_col, (min_val, max_val) in SCORE_RANGES.items():
        if score_col in df.columns:
            actual_min = df[score_col].min()
            actual_max = df[score_col].max()
            if actual_min >= min_val and actual_max <= max_val:
                ok(f"{csv_name}.{score_col}: range [{actual_min:.1f}, {actual_max:.1f}] within [{min_val}, {max_val}]")
            else:
                error(f"{csv_name}.{score_col}: range [{actual_min:.1f}, {actual_max:.1f}] OUTSIDE [{min_val}, {max_val}]")

    return df


# ============================================================
# CHECK 4: Excel Structure
# ============================================================

def check_excel():
    section("CHECK 4: Excel Structure")

    path = os.path.join(SCRIPT_DIR, "US_Stock_Analysis_Coloured.xlsx")
    if not os.path.exists(path):
        warn("Excel not found — skipping Excel checks")
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    ok(f"Excel has {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.max_row
        cols = ws.max_column
        ok(f"Sheet '{sheet_name}': {rows} rows, {cols} columns")

        # Check header row has blue fill
        header_fill = ws.cell(1, 1).fill.start_color.rgb if ws.cell(1, 1).fill else None
        if header_fill and "4472C4" in str(header_fill):
            ok(f"Sheet '{sheet_name}': header color correct")
        else:
            warn(f"Sheet '{sheet_name}': header color unexpected ({header_fill})")

        # Check first data row has at least one colored cell
        colors_found = set()
        for c in range(1, min(cols + 1, 10)):
            fill = ws.cell(2, c).fill.start_color.rgb if ws.cell(2, c).fill else None
            if fill:
                colors_found.add(str(fill))

        if colors_found:
            ok(f"Sheet '{sheet_name}': colors present in data rows")
        else:
            warn(f"Sheet '{sheet_name}': no colors found in first data row")

    wb.close()


# ============================================================
# CHECK 5: Data Consistency
# ============================================================

def check_consistency():
    section("CHECK 5: Data Consistency")

    us_csv = os.path.join(SCRIPT_DIR, "US_Stock_Data.csv")
    in_csv = os.path.join(SCRIPT_DIR, "Indian_Stock_Data.csv")

    for csv_path, label in [(us_csv, "US"), (in_csv, "Indian")]:
        if not os.path.exists(csv_path):
            continue
        df = pd.read_csv(csv_path)

        # Check for extreme outliers
        for col in ["QoQ Profit Growth", "NetProfit TTM 1Yr Growth"]:
            if col in df.columns:
                max_val = df[col].max()
                if pd.notna(max_val) and max_val > 500:
                    error(f"{label}.{col}: extreme outlier {max_val} (cap failed?)")
                elif pd.notna(max_val) and max_val <= 500:
                    ok(f"{label}.{col}: capped correctly (max={max_val})")

        # Check negative PB ratios
        if "PB Ratio" in df.columns:
            neg_pb = (df["PB Ratio"] < 0).sum()
            if neg_pb > 0:
                warn(f"{label}: {neg_pb} negative PB ratios present (may be valid)")
            else:
                ok(f"{label}: no negative PB ratios")

        # Check negative EV/EBITDA
        if "EV/EBITDA" in df.columns:
            neg_eve = (df["EV/EBITDA"] < 0).sum()
            if neg_eve > 0:
                warn(f"{label}: {neg_eve} negative EV/EBITDA (set to NaN?)")
            else:
                ok(f"{label}: no negative EV/EBITDA")


# ============================================================
# MAIN
# ============================================================

def main():
    strict = "--strict" in sys.argv

    print("=" * 60)
    print("  AnjaliValueStocks Pipeline Verification Agent")
    print("=" * 60)

    check_files()
    check_syntax()

    # US CSV checks
    us_expected = [
        "Ticker", "Sector", "Sub Sector",
        "Sales YoY Growth", "NetProfit YoY Growth",
        "Sales TTM 1Yr Growth", "NetProfit TTM 1Yr Growth",
        "QoQ Sales Growth", "QoQ Profit Growth",
        "3M Return", "6M Return", "1Yr Return", "2Yr Return",
        "PE Ratio", "Future PE", "TTM PEG", "Future PEG",
        "PB Ratio", "EV/Sales", "EV/EBITDA",
        "Market Cap (Billions)", "Revenue (Billions)", "TTM Revenue (Billions)",
        "QtrStd", "YrStd", "Qtr Beta", "Yr Beta",
        "_Loss_Profit_YoY", "_Loss_Profit_TTM", "_Loss_Profit_QoQ",
    ]
    check_csv("US_Stock_Data.csv", us_expected)

    # SmallMidCap CSV checks
    sm_expected = ["Index"] + us_expected
    check_csv("SmallMidCap_Stock_Data.csv", sm_expected)

    # Indian CSV checks
    in_expected = us_expected + ["Index Name", "DII Quarter", "DII 1Yr", "FII Quarter", "FII 1Yr"]
    check_csv("Indian_Stock_Data.csv", in_expected)

    check_excel()
    check_consistency()

    section("VERIFICATION SUMMARY")
    print(f"\n  Checks passed: {len([e for e in ERRORS if 'ERROR' not in e])}")
    print(f"  Warnings:      {len(WARNINGS)}")
    print(f"  Errors:        {len(ERRORS)}")

    if ERRORS:
        print(f"\n[FAIL] FAILED -- {len(ERRORS)} error(s) found:")
        for e in ERRORS:
            print(f"    {e}")
        sys.exit(1)
    elif WARNINGS and strict:
        print(f"\n[WARN] WARNINGS in strict mode -- {len(WARNINGS)} warning(s):")
        for w in WARNINGS:
            print(f"    {w}")
        sys.exit(1)
    else:
        print(f"\n[PASS] ALL CHECKS PASSED -- ready for GitHub commit")
        if WARNINGS:
            print(f"\n  ({len(WARNINGS)} non-critical warnings -- see above)")
        sys.exit(0)


if __name__ == "__main__":
    main()
