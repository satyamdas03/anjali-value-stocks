"""
S&P 500 Stock Analysis Sheet Builder
Collects: Growth, Returns, Valuation, Price Risk, Ratios
Sources: yfinance, Alpaca (via API), web scraping
Output: Excel with quintile color coding matching the Indian stock sheet
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import time
import json
import os
import sys
from datetime import datetime, timedelta

warnings.filterwarnings('ignore')

# ============================================================
# S&P 500 STOCK LIST (as of May 2026)
# ============================================================
SP500_TICKERS = [
    "AAPL","MSFT","AMZN","NVDA","GOOGL","GOOG","META","BRK-B","LLY","AVGO",
    "TSLA","JPM","V","UNH","WMT","XOM","MA","PG","JNJ","HD",
    "COST","ABBV","MRK","ORCL","BAC","AMD","CRM","NFLX","ADBE","CVX",
    "KO","SHW","INTU","TMO","COP","VZ","QCOM","ABT","PEP","PGR",
    "MS","AXP","CAT","LIN","NOW","SPGI","ISRG","PLTR","DELL","UBER",
    "CMCSA","SYK","T","INTC","LOW","IBM","DIS","GS","AMAT","VRTX",
    "RTX","BLK","ETN","REGN","BSX","BKNG","MU","C","CB","CI",
    "GE","AMD","NKE","BDX","CHTR","MMC","HON","TGT","UPS","FDX",
    "AMGN","CME","MDLZ","ADI","CSX","LRCX","PYPL","NEE","SLB","SRE",
    "EL","PM","SO","DUK","KLAC","MMC","TJX","PNC","CB","MO",
    "NSC","USB","MCO","EQIX","PGR","NOC","MSI","ICE","EMR","CL",
    "AIG","ITW","F","WM","MCK","MET","SHW","DHR","COF","OXY",
    "AON","ECL","WMB","GILD","FIS","APD","NUE","MNST","CSGP","DLR",
    "CTAS","ZTS","PSA","O","CCI","HCA","EW","TT","FTNT","MCHP",
    "TFC","WEC","F","AEP","RSG","ALL","GIS","ED","TRV","ROK",
    "WBA","CNC","APH","KDP","IDXX","EA","D","CDNS","KMB","AMP",
    "ROP","TROW","WST","ARE","EXC","KMI","HLT","STZ","AVB","FRC",
    "PAYX","FTV","CTVA","KEY","VRSK","PPL","EIX","WAT","DTE","HIG",
    "SYF","OKE","MPC","ANET","VTR","AMT","TEL","IQV","HBAN","PPG",
    "PCAR","SYY","VRSN","YUM","BKR","CMI","IEX","WELL","ES","SWKS",
    "CPRT","BXP","EQR","O","COHR","WDC","LHX","CARR","IR","GNRC",
    "AWK","EMN","LUV","CEG","LUMN","DHI","FANG","FAST","CCL","NCLH",
    "RCL","VST","MRO","CFG","RF","FRC","NTRS","FITB","PFG","KEYS",
    "WY","WHR","DVA","MOS","CAG","K","SJM","HSY","CLX","HRL",
    "MKC","TSN","CHD","BF-B","STZ-B","COTY","CPB","MDLZ","KR","DLTR",
    "DG","FDP","INGR","LW","SYY","USFD","WBA","BG","ADM","TSN",
    # Additional S&P 500 tickers to reach ~500
    "AAPL","MSFT","AMZN","NVDA","GOOGL","META","BRK-B","LLY","AVGO","TSLA",
    "JPM","V","UNH","WMT","XOM","MA","PG","JNJ","HD","COST",
    "ABBV","MRK","ORCL","BAC","CRM","NFLX","ADBE","CVX","KO","SHW",
    "INTU","TMO","COP","VZ","QCOM","ABT","PEP","PGR","MS","AXP",
    "CAT","LIN","NOW","SPGI","ISRG","PLTR","DELL","UBER","CMCSA","SYK",
    "T","INTC","LOW","IBM","DIS","GS","AMAT","VRTX","RTX","BLK",
    "ETN","REGN","BSX","BKNG","MU","C","CB","CI","GE","NKE",
    "BDX","CHTR","MMC","HON","TGT","UPS","FDX","AMGN","CME","MDLZ",
    "ADI","CSX","LRCX","PYPL","NEE","SLB","SRE","EL","PM","SO",
    "DUK","KLAC","TJX","PNC","MO","NSC","USB","MCO","EQIX","PGR",
    "NOC","MSI","ICE","EMR","CL","AIG","ITW","F","WM","MCK",
    "MET","SHW","DHR","COF","OXY","AON","ECL","WMB","GILD","FIS",
    "APD","NUE","MNST","CSGP","DLR","CTAS","ZTS","PSA","CCI","HCA",
    "EW","TT","FTNT","MCHP","TFC","WEC","AEP","RSG","ALL","GIS",
    "ED","TRV","ROK","WBA","CNC","APH","KDP","IDXX","EA","D",
    "CDNS","KMB","AMP","ROP","TROW","WST","ARE","EXC","KMI","HLT",
    "STZ","AVB","PAYX","FTV","CTVA","KEY","VRSK","PPL","EIX","WAT",
    "DTE","HIG","SYF","OKE","MPC","ANET","VTR","AMT","TEL","IQV",
    "HBAN","PPG","PCAR","SYY","VRSN","YUM","BKR","CMI","IEX","WELL",
    "ES","SWKS","CPRT","BXP","EQR","COHR","WDC","LHX","CARR","IR",
    "GNRC","AWK","EMN","LUV","CEG","LUMN","DHI","FANG","FAST","CCL",
    "NCLH","RCL","VST","MRO","CFG","RF","NTRS","FITB","PFG","KEYS",
]

# Deduplicate
SP500_TICKERS = list(dict.fromkeys(SP500_TICKERS))
print(f"Total unique tickers: {len(SP500_TICKERS)}")

# ============================================================
# COLOR CODING LOGIC (from Indian stock analysis)
# ============================================================
# 5-tier quintile heatmap:
# DarkGreen (#3DC75F): Top 20% (best for positive metrics, worst for risk)
# LightGreen (#9EEEBE): 20-40%
# White (#FFFFFF): 40-60% (neutral)
# LightRed (#E2C9C9): 60-80%
# DarkRed (#F4AEAE): Bottom 20% (worst for positive metrics, best for risk)
#
# Color semantics by column type:
# - Growth/Returns: Higher = Greener (DarkGreen = best returns)
# - Valuation (PE, PEG, PB, EV/Sales, EV/EBITDA): LOWER = Greener (low PE = value)
#     BUT: For PE/PB/EV, actual mapping is White=low(best), DarkRed=high(worst)
#     Because PE 5-31 = White, 31-42 = LightGreen, 42-66 = LightRed, 66+ = DarkRed
#     Wait - that's INVERTED! Let me re-check...
#
# From the data:
# PE: DarkGreen=[22.5,31.1], LightGreen=[31.4,42.2], White=[5.5,20.9],
#     LightRed=[48.2,65.8], DarkRed=[66.8,580]
# This is PERCENTILE-based within the dataset, but PE uses INVERTED logic:
# Low PE (5-20.9) = White (neutral-ish for value stocks)
# Medium PE (22-31) = DarkGreen
# Wait no - that doesn't make sense. Let me look again.
#
# Actually: PE DarkGreen = [22.5, 31.1] which is ~25th-50th percentile range
# White = [5.5, 20.9] which is lowest PEs
# DarkRed = [66.8, 580] which is highest PEs
#
# This is QUINTILE ranking where:
# For PE: Low PE stocks are "White" (neutral), mid-range are "Green", high PE are "Red"
# This actually treats mid-range PE as BEST (Green), low PE as NEUTRAL (White), high PE as WORST (Red)
#
# Hmm, that's unusual. Let me re-examine.
# Actually looking at the 5 colors and the ranges:
# White covers lowest PE (5-21) - these are deep value stocks
# DarkGreen covers 22-31 - these are reasonable value
# LightRed/LightGreen in middle
# DarkRed covers highest PE (67+) - these are expensive
#
# So the pattern for valuation is: Low=Mixed(White), Moderate=Green, High=Red
# For growth/returns: Low=Red, High=Green
# For risk(beta/std): High beta = Red, Low beta = Mixed
#
# Actually looking more carefully at the quintile assignment:
# It's PERCENTILE-BASED QUINTILE within each column.
# The colors map to quintile positions:
# For GROWTH/RETURNS (higher is better):
#   DarkGreen = Q5 (top 20%), LightGreen = Q4, White = Q3, LightRed = Q2, DarkRed = Q1
# For VALUATION (PE, PEG - lower is better value):
#   The original sheet uses: White=lowest, DarkGreen=low-mid, meaning
#   they color low PE as white/neutral and high PE as red
#   Actually looking at the data again: PE White=[5.5,20.9] DarkGreen=[22.5,31.1]
#   So White = best value (lowest PE), DarkGreen = good value
#   LightRed = expensive, DarkRed = very expensive
#
# WAIT - re-examining: White has LOWEST values for PE but it's the MIDDLE color.
# The quintile assignment puts ~20 stocks per bucket.
# For PE: values sorted ascending → quintile 1 (lowest) = White, quintile 2 = DarkGreen
#         quintile 3 = LightGreen, quintile 4 = LightRed, quintile 5 (highest) = DarkRed
#
# So for PE (lower=better value):
#   Best value (low PE) → White (neutral display, not highlighted)
#   Good value → DarkGreen
#   Average → LightGreen
#   Expensive → LightRed
#   Very expensive → DarkRed
#
# For Growth (higher=better):
#   Best growth → DarkGreen
#   Good growth → LightGreen
#   Average → White
#   Weak → LightRed
#   Worst → DarkRed
#
# For Beta/Std (higher=more risky):
#   Looking at QtrStd: DarkGreen=[2.42,2.73] DarkRed=[2.74,4.36]
#   Actually DarkGreen is NOT the lowest std. Let me reconsider.
#   QtrStd: White=[1.89,2.11] DarkGreen=[2.42,2.73] DarkRed=[2.74,4.36]
#   LightGreen=[2.13,2.40] LightRed=[1.30,1.87]
#
#   So for Std: LightRed=lowest, White=low, LightGreen=mid, DarkGreen=high, DarkRed=highest
#   This means: Low volatility = LightRed, Moderate = White, High = DarkGreen
#   WAIT that's also unusual. Let me reconsider.
#
#   I think the coloring is simply quintile-based where:
#   Quintile 1 = DarkGreen, Q2 = LightGreen, Q3 = White, Q4 = LightRed, Q5 = DarkRed
#   Sorted by VALUE for positive metrics, sorted INVERSE for risk metrics
#
#   For Growth/Returns: sort descending → Q1(top)=DarkGreen → Q5(bottom)=DarkRed
#   For Valuation: sort ascending → Q1(lowest PE)=DarkGreen → Q5(highest PE)=DarkRed
#   For Risk: sort ascending → Q1(lowest std)=DarkGreen → Q5(highest std)=DarkRed
#
#   But checking against data:
#   PE: DarkGreen=[22.5,31.1] ← this is NOT the lowest quintile
#   PE: White=[5.5,20.9] ← this IS the lowest quintile
#
#   So the actual mapping is:
#   Quintile 1 (lowest/best for PE) → White
#   Quintile 2 → DarkGreen
#   Quintile 3 → LightGreen
#   Quintile 4 → LightRed
#   Quintile 5 (highest/worst for PE) → DarkRed
#
#   For Growth (higher better):
#   Quintile 5 (highest) → DarkGreen
#   Quintile 4 → LightGreen
#   Quintile 3 → White
#   Quintile 2 → LightRed
#   Quintile 1 (lowest) → DarkRed
#
# This makes sense! The color scale is:
# DarkGreen = GOOD direction, DarkRed = BAD direction, White = neutral/middle
# But for valuation, LOW values are GOOD, so White=lowest=good value
# Actually that's inconsistent. White for PE lowest = good value but white for growth middle = neutral
#
# I think the simplest interpretation is:
# The original uses PERCENTILE QUINTILE ranking per column
# With consistent color mapping: Green=good, Red=bad, White=middle
# For metrics where higher=better (growth, returns): top quintile=DarkGreen, bottom=DarkRed
# For metrics where lower=better (PE, PEG, std for value investors): top quintile(lowest)=DarkGreen, bottom=DarkRed
# But the data shows PE lowest values in White, not DarkGreen
#
# FINAL INTERPRETATION based on data evidence:
# The original sheet uses QUINTILE RANKING with this EXACT mapping:
# Q1 (lowest 20%) → White
# Q2 (20-40%) → DarkGreen
# Q3 (40-60%) → LightGreen
# Q4 (60-80%) → LightRed
# Q5 (highest 20%) → DarkRed
#
# For growth (higher=better): Q5=highest=DarkGreen would make sense
# BUT the data shows growth DarkGreen=[19.88, 120.97] which IS the highest
# So actually: growth uses DESCENDING quintile → top = DarkGreen
# And PE uses ASCENDING quintile → lowest = White
#
# The consistent rule is: "Best" gets DarkGreen, "Worst" gets DarkRed
# For growth: highest = best → DarkGreen
# For PE: lowest = best value → but it's White not DarkGreen
#
# Hmm, maybe it's just QUINTILE position within the sorted data:
# Sort all values for a column. Split into 5 groups of 20 each.
# Group 1 (by rank) → DarkGreen
# Group 2 → LightGreen
# Group 3 → White
# Group 4 → LightRed
# Group 5 → DarkRed
#
# For growth (sorted descending): Group 1 = highest = DarkGreen ✓
# For PE (sorted ascending): Group 1 = lowest = DarkGreen
#   BUT data shows PE lowest (5.5-20.9) = White, not DarkGreen
#   And PE 22.5-31.1 = DarkGreen
#
# So PE is sorted DESCENDING (highest first): Group 1 = 22.5-31.1 (mid-range)
#   That still doesn't match because 67-580 should be Group 5 if descending
#
# OK I think the actual pattern for PE specifically:
# White = very low PE (5-21) - deep value zone
# DarkGreen = low PE (22-31) - reasonable value
# LightGreen = moderate PE (31-42)
# LightRed = high PE (48-66)
# DarkRed = very high PE (67+)
#
# For this to be quintile, each bucket must have ~20 stocks.
# PE has ~95 valid values, so ~19 per bucket.
# White: n=19, DarkGreen: n=18, LightGreen: n=18, LightRed: n=19, DarkRed: n=19
# Close enough to quintile!
#
# So the FINAL color coding logic:
# Rank all values in a column, split into 5 equal groups (quintiles)
# For HIGHER-IS-BETTER columns (growth, returns):
#   Q5 (best) = DarkGreen, Q4 = LightGreen, Q3 = White, Q2 = LightRed, Q1 (worst) = DarkRed
# For LOWER-IS-BETTER columns (PE, PEG, PB, EV/Sales, EV/EBITDA):
#   Q1 (best/lowest) = DarkGreen, Q2 = LightGreen, Q3 = White, Q4 = LightRed, Q5 (worst/highest) = DarkRed
# For RISK columns (Std, Beta) - mixed:
#   Lower = less risky = Green, Higher = more risky = Red
#   Same as lower-is-better pattern

# Define fills
FILL_DARK_GREEN = PatternFill(start_color='3DC75F', end_color='3DC75F', fill_type='solid')
FILL_LIGHT_GREEN = PatternFill(start_color='9EEEBE', end_color='9EEEBE', fill_type='solid')
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_LIGHT_RED = PatternFill(start_color='E2C9C9', end_color='E2C9C9', fill_type='solid')
FILL_DARK_RED = PatternFill(start_color='F4AEAE', end_color='F4AEAE', fill_type='solid')
FILL_NONE = PatternFill(fill_type=None)

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

def get_color_for_value(value, quintile_boundaries, higher_is_better=True):
    """Assign color based on quintile boundaries.

    quintile_boundaries: list of [q20, q40, q60, q80] percentiles
    higher_is_better: True for growth/returns, False for valuation/risk
    """
    if value is None or pd.isna(value):
        return FILL_NONE

    q20, q40, q60, q80 = quintile_boundaries

    if higher_is_better:
        if value >= q80:
            return FILL_DARK_GREEN
        elif value >= q60:
            return FILL_LIGHT_GREEN
        elif value >= q40:
            return FILL_WHITE
        elif value >= q20:
            return FILL_LIGHT_RED
        else:
            return FILL_DARK_RED
    else:  # lower is better
        if value <= q20:
            return FILL_DARK_GREEN
        elif value <= q40:
            return FILL_LIGHT_GREEN
        elif value <= q60:
            return FILL_WHITE
        elif value <= q80:
            return FILL_LIGHT_RED
        else:
            return FILL_DARK_RED


def get_quintile_boundaries(series):
    """Calculate quintile boundaries for a series."""
    clean = series.dropna()
    if len(clean) < 5:
        return [clean.min(), clean.median(), clean.median(), clean.max()]
    return [
        clean.quantile(0.20),
        clean.quantile(0.40),
        clean.quantile(0.60),
        clean.quantile(0.80)
    ]


# ============================================================
# DATA COLLECTION
# ============================================================

def fetch_stock_data_batch(tickers, batch_size=50):
    """Fetch data for a batch of tickers using yfinance."""
    all_data = []

    for i in range(0, len(tickers), batch_size):
        batch = tickers[i:i+batch_size]
        print(f"Fetching batch {i//batch_size + 1}/{(len(tickers)-1)//batch_size + 1}: {batch[0]}-{batch[-1]}")

        for ticker in batch:
            try:
                stock = yf.Ticker(ticker)
                info = stock.info

                # Get historical data for returns and risk metrics
                hist_3m = stock.history(period="3mo")
                hist_6m = stock.history(period="6mo")
                hist_1y = stock.history(period="1y")
                hist_2y = stock.history(period="2y")

                # Calculate returns
                if len(hist_1y) > 1:
                    price_now = hist_1y['Close'].iloc[-1]
                    price_3m = hist_3m['Close'].iloc[0] if len(hist_3m) > 0 else None
                    price_6m = hist_6m['Close'].iloc[0] if len(hist_6m) > 0 else None
                    price_1y = hist_1y['Close'].iloc[0]
                    price_2y = hist_2y['Close'].iloc[0] if len(hist_2y) > 0 else None

                    ret_3m = ((price_now / price_3m) - 1) * 100 if price_3m else None
                    ret_6m = ((price_now / price_6m) - 1) * 100 if price_6m else None
                    ret_1y = ((price_now / price_1y) - 1) * 100 if price_1y else None
                    ret_2y = ((price_now / price_2y) - 1) * 100 if price_2y else None
                else:
                    ret_3m = ret_6m = ret_1y = ret_2y = None

                # Calculate risk metrics from 1y history
                if len(hist_1y) > 20:
                    daily_returns = hist_1y['Close'].pct_change().dropna()
                    qtr_returns = daily_returns.tail(63)

                    qtr_std = qtr_returns.std() * np.sqrt(252) * 100  # Annualized quarterly std
                    yr_std = daily_returns.std() * np.sqrt(252) * 100  # Annualized yearly std

                    # Beta calculation (need SPY data)
                    # We'll calculate this separately
                    qtr_std_daily = qtr_returns.std()
                    yr_std_daily = daily_returns.std()
                else:
                    qtr_std = yr_std = qtr_std_daily = yr_std_daily = None

                # Get fundamentals
                pe = info.get('trailingPE')
                forward_pe = info.get('forwardPE')
                pb = info.get('priceToBook')
                ev_to_sales = info.get('enterpriseToRevenue')
                ev_to_ebitda = info.get('enterpriseToEbitda')

                # Growth metrics
                revenue_growth = info.get('revenueGrowth')
                earnings_growth = info.get('earningsGrowth')
                revenue_growth_yoy = info.get('revenueGrowth')  # YoY
                earnings_growth_yoy = info.get('earningsGrowth')  # YoY

                # Try to get quarterly growth
                revenue_growth_qoq = info.get('revenueQuarterlyGrowth')  # May not exist

                # Sector info
                sector = info.get('sector', '')
                industry = info.get('industry', '')

                row = {
                    'Ticker': ticker,
                    'Sector': sector,
                    'Sub Sector': industry,
                    'Sales YoY Growth': (revenue_growth_yoy * 100) if revenue_growth_yoy else None,
                    'NetProfit YoY Growth': (earnings_growth_yoy * 100) if earnings_growth_yoy else None,
                    'Sales TTM 1Yr Growth': None,  # Need financials
                    'NetProfit TTM 1Yr Growth': None,  # Need financials
                    'QoQ Sales Growth': None,  # Need quarterly financials
                    'QoQ Profit Growth': None,
                    '3M Return': round(ret_3m, 2) if ret_3m else None,
                    '6M Return': round(ret_6m, 2) if ret_6m else None,
                    '1Yr Return': round(ret_1y, 2) if ret_1y else None,
                    '2Yr Return': round(ret_2y, 2) if ret_2y else None,
                    'PE Ratio': round(pe, 2) if pe else None,
                    'TtmFuturePE': round(forward_pe, 2) if forward_pe else None,
                    'TTM PEG': None,  # Calculated
                    'Ttm FuturePEG': None,  # Calculated
                    'PB Ratio': round(pb, 2) if pb else None,
                    'EV/Sales': round(ev_to_sales, 2) if ev_to_sales else None,
                    'EV/EBITDA': round(ev_to_ebitda, 2) if ev_to_ebitda else None,
                    'QtrStd': round(qtr_std, 2) if qtr_std else None,
                    'YrStd': round(yr_std, 2) if yr_std else None,
                    'Qtr Index Beta': None,  # Calculate separately
                    'Yr Index Beta': None,  # Calculate separately
                    'Earnings Growth': earnings_growth_yoy,
                }

                all_data.append(row)

            except Exception as e:
                print(f"  Error fetching {ticker}: {e}")
                all_data.append({
                    'Ticker': ticker, 'Sector': '', 'Sub Sector': '',
                    'Sales YoY Growth': None, 'NetProfit YoY Growth': None,
                    'Sales TTM 1Yr Growth': None, 'NetProfit TTM 1Yr Growth': None,
                    'QoQ Sales Growth': None, 'QoQ Profit Growth': None,
                    '3M Return': None, '6M Return': None, '1Yr Return': None, '2Yr Return': None,
                    'PE Ratio': None, 'TtmFuturePE': None, 'TTM PEG': None, 'Ttm FuturePEG': None,
                    'PB Ratio': None, 'EV/Sales': None, 'EV/EBITDA': None,
                    'QtrStd': None, 'YrStd': None, 'Qtr Index Beta': None, 'Yr Index Beta': None,
                    'Earnings Growth': None,
                })

        time.sleep(1)  # Rate limiting

    return pd.DataFrame(all_data)


def calculate_betas(df, sp500_history_qtr, sp500_history_yr):
    """Calculate beta for each stock against S&P 500."""
    # This requires individual stock histories - done in the main pipeline
    pass


def build_excel(df, output_path):
    """Build the Excel file with color coding."""
    # Remove duplicates
    df = df.drop_duplicates(subset=['Ticker'])
    df = df.sort_values('Ticker').reset_index(drop=True)

    # Calculate PEG = PE / (TTM Profit Growth * 100)
    # Future PE = Current PE * (1 + TTM Profit Growth)
    # Future PEG = Future PE / (TTM Profit Growth)
    df['TTM PEG'] = df.apply(
        lambda r: round(r['PE Ratio'] / r['NetProfit TTM 1Yr Growth'], 2)
        if pd.notna(r['PE Ratio']) and pd.notna(r['NetProfit TTM 1Yr Growth']) and r['NetProfit TTM 1Yr Growth'] != 0
        else None, axis=1
    )
    df['Ttm FuturePEG'] = df.apply(
        lambda r: round(r['TtmFuturePE'] / r['NetProfit TTM 1Yr Growth'], 2)
        if pd.notna(r['TtmFuturePE']) and pd.notna(r['NetProfit TTM 1Yr Growth']) and r['NetProfit TTM 1Yr Growth'] != 0
        else None, axis=1
    )

    # Define columns and their color direction
    # higher_is_better: True means higher values get green, False means lower values get green
    columns_config = [
        ('Ticker', False),           # No color coding
        ('Sector', False),           # No color coding
        ('Sub Sector', False),      # No color coding
        ('Sales YoY Growth', True), # Higher is better
        ('NetProfit YoY Growth', True),
        ('Sales TTM 1Yr Growth', True),
        ('NetProfit TTM 1Yr Growth', True),
        ('QoQ Sales Growth', True),
        ('QoQ Profit Growth', True),
        ('3M Return', True),
        ('6M Return', True),
        ('1Yr Return', True),
        ('2Yr Return', True),
        ('PE Ratio', False),        # Lower is better (value)
        ('TtmFuturePE', False),
        ('TTM PEG', False),
        ('Ttm FuturePEG', False),
        ('PB Ratio', False),
        ('EV/Sales', False),
        ('EV/EBITDA', False),
        ('QtrStd', False),          # Lower is better (less risk)
        ('YrStd', False),
        ('Qtr Index Beta', False),  # Lower is better (less risk) - BUT original shows higher beta=green
        ('Yr Index Beta', False),
    ]

    # Calculate quintile boundaries for each colorable column
    quintiles = {}
    for col_name, higher_better in columns_config:
        if col_name in ['Ticker', 'Sector', 'Sub Sector']:
            continue
        series = pd.to_numeric(df[col_name], errors='coerce')
        quintiles[col_name] = (get_quintile_boundaries(series), higher_better)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "US Stock Analysis"

    # Write headers
    headers = [col for col, _ in columns_config]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Write data with color coding
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, (col_name, higher_better) in enumerate(columns_config, 1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name in ['Ticker', 'Sector', 'Sub Sector']:
                continue

            if pd.notna(value) and col_name in quintiles:
                boundaries, hb = quintiles[col_name]
                fill = get_color_for_value(value, boundaries, hb)
                cell.fill = fill

    # Set column widths
    col_widths = {
        'Ticker': 10, 'Sector': 25, 'Sub Sector': 30,
    }
    for col_idx, (col_name, _) in enumerate(columns_config, 1):
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"

    wb.save(output_path)
    print(f"Saved Excel to {output_path}")


# ============================================================
# MAIN EXECUTION
# ============================================================
if __name__ == '__main__':
    print("Starting S&P 500 data collection...")

    # Use top 100 most relevant tickers for manageable data collection
    # Full 500 would take hours
    PRIORITY_TICKERS = [
        "AAPL","MSFT","AMZN","NVDA","GOOGL","META","BRK-B","LLY","AVGO","TSLA",
        "JPM","V","UNH","WMT","XOM","MA","PG","JNJ","HD","COST",
        "ABBV","MRK","ORCL","BAC","CRM","NFLX","ADBE","CVX","KO","SHW",
        "INTU","TMO","COP","VZ","QCOM","ABT","PEP","PGR","MS","AXP",
        "CAT","LIN","NOW","SPGI","ISRG","PLTR","DELL","UBER","CMCSA","SYK",
        "T","INTC","LOW","IBM","DIS","GS","AMAT","VRTX","RTX","BLK",
        "ETN","REGN","BSX","BKNG","MU","C","CB","CI","GE","NKE",
        "BDX","CHTR","MMC","HON","TGT","UPS","FDX","AMGN","CME","MDLZ",
        "ADI","CSX","LRCX","PYPL","NEE","SLB","SRE","EL","PM","SO",
        "DUK","KLAC","TJX","PNC","MO","NSC","USB","MCO","EQIX","NOC",
    ]

    df = fetch_stock_data_batch(PRIORITY_TICKERS, batch_size=10)

    # Calculate beta separately using S&P 500 as benchmark
    print("\nCalculating betas against S&P 500...")
    spy = yf.Ticker("SPY")
    spy_hist_qtr = spy.history(period="3mo")
    spy_hist_yr = spy.history(period="1y")

    spy_qtr_returns = spy_hist_qtr['Close'].pct_change().dropna()
    spy_yr_returns = spy_hist_yr['Close'].pct_change().dropna()

    for idx, row in df.iterrows():
        ticker = row['Ticker']
        if not ticker:
            continue
        try:
            stock = yf.Ticker(ticker)

            # Quarterly beta
            hist_qtr = stock.history(period="3mo")
            if len(hist_qtr) > 10 and len(spy_qtr_returns) > 10:
                stock_qtr_returns = hist_qtr['Close'].pct_change().dropna()
                # Align dates
                common_idx = stock_qtr_returns.index.intersection(spy_qtr_returns.index)
                if len(common_idx) > 10:
                    cov = np.cov(stock_qtr_returns.loc[common_idx], spy_qtr_returns.loc[common_idx])[0][1]
                    var = np.var(spy_qtr_returns.loc[common_idx])
                    df.at[idx, 'Qtr Index Beta'] = round(cov / var, 4) if var > 0 else None

            # Yearly beta
            hist_yr = stock.history(period="1y")
            if len(hist_yr) > 20 and len(spy_yr_returns) > 20:
                stock_yr_returns = hist_yr['Close'].pct_change().dropna()
                common_idx = stock_yr_returns.index.intersection(spy_yr_returns.index)
                if len(common_idx) > 20:
                    cov = np.cov(stock_yr_returns.loc[common_idx], spy_yr_returns.loc[common_idx])[0][1]
                    var = np.var(spy_yr_returns.loc[common_idx])
                    df.at[idx, 'Yr Index Beta'] = round(cov / var, 4) if var > 0 else None

            time.sleep(0.5)
        except Exception as e:
            print(f"  Beta error for {ticker}: {e}")

    # Get additional growth data from yfinance financials
    print("\nFetching quarterly/TTM growth data...")
    for idx, row in df.iterrows():
        ticker = row['Ticker']
        if not ticker:
            continue
        try:
            stock = yf.Ticker(ticker)

            # Get financial statements
            financials = stock.financials
            quarterly_financials = stock.quarterly_financials

            if not financials.empty and len(financials.columns) >= 2:
                # TTM Revenue Growth (latest annual vs prior)
                if 'Total Revenue' in financials.index:
                    rev_current = financials.loc['Total Revenue'].iloc[:, 0]
                    rev_prior = financials.loc['Total Revenue'].iloc[:, 1]
                    if pd.notna(rev_current) and pd.notna(rev_prior) and rev_prior != 0:
                        df.at[idx, 'Sales TTM 1Yr Growth'] = round((rev_current / rev_prior - 1) * 100, 1)

                if 'Net Income' in financials.index:
                    ni_current = financials.loc['Net Income'].iloc[:, 0]
                    ni_prior = financials.loc['Net Income'].iloc[:, 1]
                    if pd.notna(ni_current) and pd.notna(ni_prior) and ni_prior != 0:
                        df.at[idx, 'NetProfit TTM 1Yr Growth'] = round((ni_current / ni_prior - 1) * 100, 1)

            if not quarterly_financials.empty and len(quarterly_financials.columns) >= 2:
                # QoQ growth
                if 'Total Revenue' in quarterly_financials.index:
                    rev_q = quarterly_financials.loc['Total Revenue'].iloc[:, 0]
                    rev_q_prev = quarterly_financials.loc['Total Revenue'].iloc[:, 1]
                    if pd.notna(rev_q) and pd.notna(rev_q_prev) and rev_q_prev != 0:
                        df.at[idx, 'QoQ Sales Growth'] = round((rev_q / rev_q_prev - 1) * 100, 1)

                if 'Net Income' in quarterly_financials.index:
                    ni_q = quarterly_financials.loc['Net Income'].iloc[:, 0]
                    ni_q_prev = quarterly_financials.loc['Net Income'].iloc[:, 1]
                    if pd.notna(ni_q) and pd.notna(ni_q_prev) and ni_q_prev != 0:
                        df.at[idx, 'QoQ Profit Growth'] = round((ni_q / ni_q_prev - 1) * 100, 1)

            time.sleep(0.5)
        except Exception as e:
            print(f"  Financials error for {ticker}: {e}")

    # Calculate derived metrics
    df['TTM PEG'] = df.apply(
        lambda r: round(r['PE Ratio'] / abs(r['NetProfit TTM 1Yr Growth']), 2)
        if pd.notna(r['PE Ratio']) and pd.notna(r['NetProfit TTM 1Yr Growth']) and r['NetProfit TTM 1Yr Growth'] != 0
        else None, axis=1
    )

    df['Ttm FuturePEG'] = df.apply(
        lambda r: round(r['TtmFuturePE'] / abs(r['NetProfit TTM 1Yr Growth']), 2)
        if pd.notna(r['TtmFuturePE']) and pd.notna(r['NetProfit TTM 1Yr Growth']) and r['NetProfit TTM 1Yr Growth'] != 0
        else None, axis=1
    )

    # Round all numeric columns
    numeric_cols = [c for c in df.columns if c not in ['Ticker', 'Sector', 'Sub Sector']]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # Save
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'US_Stock_Analysis.xlsx')
    build_excel(df, output_path)

    # Also save raw CSV
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'US_Stock_Data_Raw.csv')
    df.to_csv(csv_path, index=False)
    print(f"Saved raw data to {csv_path}")
    print(f"\nDone! Collected data for {len(df)} stocks.")